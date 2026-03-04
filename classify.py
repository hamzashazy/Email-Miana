"""
Classify emails from an xlsx file or a single email via CLI.

Usage:
    # Classify all emails from the dataset
    python classify.py --input email-classification-report.xlsx --output results.xlsx

    # Classify a single email
    python classify.py --single --subject "Re: Cash Offer for 123 Main St" --body "No thanks, not interested"
"""

import argparse
import asyncio
import json
import logging
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from classifier.config import MAX_CONCURRENT_REQUESTS
from classifier.data_loader import load_emails
from classifier.llm_client import classify_batch, classify_email

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger(__name__)

CATEGORY_FILLS = {
    "rejection": PatternFill(start_color="FFD9D9", end_color="FFD9D9", fill_type="solid"),
    "counter_offer": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
    "interested": PatternFill(start_color="D9F2D9", end_color="D9F2D9", fill_type="solid"),
    "replies": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
    "needs_review": PatternFill(start_color="E8D5F5", end_color="E8D5F5", fill_type="solid"),
}


def write_results_xlsx(results: list[dict], output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classification Results"

    headers = [
        "ID", "LLM Classification", "Confidence", "Reasoning",
        "Original Classification", "Match", "Subject", "Snippet",
        "From", "Property Address", "Offer Type", "Offer Price", "Received At",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, r in enumerate(results, 2):
        original = r.get("classification_original", r.get("classification", ""))
        llm_cls = r.get("classification", "")
        match = "Y" if original == llm_cls else "N"

        values = [
            r.get("id", ""),
            llm_cls,
            f"{r.get('confidence', 0):.0%}",
            r.get("reasoning", ""),
            original,
            match,
            r.get("subject", ""),
            (r.get("snippet", "") or "")[:200],
            r.get("from_email", ""),
            r.get("property_address", ""),
            r.get("offer_type", ""),
            r.get("offer_price", ""),
            str(r.get("received_at", "")),
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)

        fill = CATEGORY_FILLS.get(llm_cls)
        if fill:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col).fill = fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 18

    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["G"].width = 45
    ws.column_dimensions["H"].width = 55

    wb.save(output_path)
    logger.info("Results written to %s", output_path)


async def run_batch(input_path: str, output_path: str, concurrency: int):
    emails = load_emails(input_path)
    logger.info("Loaded %d emails", len(emails))

    original_labels = {e["id"]: e["classification"] for e in emails}

    start = time.time()
    results = await classify_batch(emails, max_concurrent=concurrency)
    elapsed = time.time() - start

    for r in results:
        r["classification_original"] = original_labels.get(r["id"], "")

    logger.info("Classified %d emails in %.1fs", len(results), elapsed)

    if output_path.endswith(".json"):
        Path(output_path).write_text(json.dumps(
            [{"id": r["id"], "classification": r["classification"],
              "confidence": r["confidence"], "reasoning": r["reasoning"]}
             for r in results],
            indent=2,
        ))
    else:
        write_results_xlsx(results, output_path)


async def run_single(subject: str, body: str):
    result = await classify_email(subject=subject, snippet=body)
    print(json.dumps(result, indent=2))


async def main():
    parser = argparse.ArgumentParser(description="Classify real estate offer reply emails")
    parser.add_argument("--input", type=str, help="Input xlsx file path")
    parser.add_argument("--output", type=str, default="classification-results.xlsx", help="Output file path (.xlsx or .json)")
    parser.add_argument("--concurrency", type=int, default=MAX_CONCURRENT_REQUESTS)
    parser.add_argument("--single", action="store_true", help="Classify a single email")
    parser.add_argument("--subject", type=str, default="", help="Email subject (for --single)")
    parser.add_argument("--body", type=str, default="", help="Email body (for --single)")
    args = parser.parse_args()

    if args.single:
        await run_single(args.subject, args.body)
    elif args.input:
        await run_batch(args.input, args.output, args.concurrency)
    else:
        parser.print_help()


if __name__ == "__main__":
    asyncio.run(main())
